#!/usr/bin/env python3
"""
UDS Diagnostic Configuration Generator
======================================
Generate UDS diagnostic code from Excel configuration
Similar to DBC generator, but for diagnostic services

Author: [小昭debug]
Date: 2026-03-15

Usage:
    python3 uds_generator.py <config.xlsx> [output_dir]
    python3 uds_generator.py --example  # Generate example config
"""

import sys
import os
import json
from datetime import datetime
from typing import Dict, List, Any, Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

# =============================================================================
# Configuration Templates
# =============================================================================

UDS_SERVICES = {
    0x10: "DiagnosticSessionControl",
    0x11: "ECUReset",
    0x14: "ClearDiagnosticInformation",
    0x19: "ReadDTCInformation",
    0x22: "ReadDataByIdentifier",
    0x27: "SecurityAccess",
    0x2E: "WriteDataByIdentifier",
    0x31: "RoutineControl",
    0x34: "RequestDownload",
    0x35: "RequestUpload",
    0x36: "TransferData",
    0x37: "RequestTransferExit",
    0x3E: "TesterPresent",
    0x85: "ControlDTCSetting",
    0x86: "ResponseOnEvent",
}

SESSION_TYPES = {
    0x01: "DefaultSession",
    0x02: "ProgrammingSession",
    0x03: "ExtendedDiagnosticSession",
    0x04: "SafetySystemDiagnosticSession",
}

SECURITY_LEVELS = {
    0x00: "Locked",
    0x01: "Level1_Seed",
    0x02: "Level1_Key",
    0x03: "Level2_Seed",
    0x04: "Level2_Key",
}

# =============================================================================
# Excel Template Generator
# =============================================================================

def create_example_config(output_path: str):
    """Create example UDS configuration Excel file"""
    
    wb = openpyxl.Workbook()
    
    # Sheet 1: Session Configuration (10服务配置)
    ws_session = wb.active
    ws_session.title = "SessionConfig"
    
    headers = ["SessionType", "Name", "P2_max(ms)", "P2*_max(ms)", 
               "S3_client(ms)", "S3_server(ms)", "SupportServices", "Description"]
    ws_session.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        cell = ws_session.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Example data for 10 service
    session_data = [
        [0x01, "DefaultSession", 50, 5000, None, 5000, "10,11,19,22,3E", "默认会话"],
        [0x02, "ProgrammingSession", 50, 5000, 2000, 5000, "10,11,27,31,34,35,36,37", "编程会话"],
        [0x03, "ExtendedSession", 50, 5000, 2000, 5000, "10,11,19,22,2E,27,31,3E", "扩展会话"],
    ]
    for row in session_data:
        ws_session.append(row)
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws_session.column_dimensions[get_column_letter(col)].width = 18
    
    # Sheet 2: DID Configuration (22/2E服务配置)
    ws_did = wb.create_sheet("DIDConfig")
    
    did_headers = ["DID(Hex)", "Name", "DataLength", "ReadOnly", 
                   "SecurityLevel", "SessionMask", "DataType", "Description"]
    ws_did.append(did_headers)
    
    for col in range(1, len(did_headers) + 1):
        cell = ws_did.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    did_data = [
        ["0xF190", "VIN", 17, "Yes", 0x01, "0x07", "ASCII", "车辆识别号"],
        ["0xF197", "ECU_Name", 16, "Yes", 0x01, "0x07", "ASCII", "ECU名称"],
        ["0xF100", "SystemState", 4, "No", 0x03, "0x06", "HEX", "系统状态"],
        ["0x0100", "VehicleSpeed", 2, "No", 0x01, "0x07", "UINT16", "车速"],
        ["0x0105", "EngineTemp", 1, "No", 0x01, "0x07", "INT8", "发动机温度"],
        ["0xF193", "SupplierCode", 4, "Yes", 0x01, "0x07", "ASCII", "供应商代码"],
        ["0xF194", "SW_Version", 8, "Yes", 0x01, "0x07", "ASCII", "软件版本"],
        ["0xF195", "HW_Version", 8, "Yes", 0x01, "0x07", "ASCII", "硬件版本"],
    ]
    for row in did_data:
        ws_did.append(row)
    
    for col in range(1, len(did_headers) + 1):
        ws_did.column_dimensions[get_column_letter(col)].width = 18
    
    # Sheet 3: Security Access (27服务配置)
    ws_sec = wb.create_sheet("SecurityConfig")
    
    sec_headers = ["Level", "Name", "SeedLength", "KeyLength", 
                   "MaxAttempts", "DelayTime(ms)", "Algorithm", "Description"]
    ws_sec.append(sec_headers)
    
    for col in range(1, len(sec_headers) + 1):
        cell = ws_sec.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    sec_data = [
        [1, "Level1_Unlock", 4, 4, 3, 10000, "XOR_Seed", "一级解锁"],
        [2, "Level2_Unlock", 4, 4, 3, 10000, "AES128", "二级解锁-编程"],
    ]
    for row in sec_data:
        ws_sec.append(row)
    
    for col in range(1, len(sec_headers) + 1):
        ws_sec.column_dimensions[get_column_letter(col)].width = 18
    
    # Sheet 4: Routine Control (31服务配置)
    ws_routine = wb.create_sheet("RoutineConfig")
    
    routine_headers = ["RID(Hex)", "Name", "StartRoutine", "StopRoutine", 
                       "GetResult", "SecurityLevel", "SessionMask", "Description"]
    ws_routine.append(routine_headers)
    
    for col in range(1, len(routine_headers) + 1):
        cell = ws_routine.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    routine_data = [
        ["0xFF01", "EraseMemory", "Yes", "No", "Yes", 0x02, "0x02", "擦除内存"],
        ["0xFF02", "CheckProgramming", "Yes", "No", "Yes", 0x02, "0x02", "校验编程"],
        ["0x0203", "SelfTest", "Yes", "No", "Yes", 0x01, "0x07", "自检例程"],
    ]
    for row in routine_data:
        ws_routine.append(row)
    
    for col in range(1, len(routine_headers) + 1):
        ws_routine.column_dimensions[get_column_letter(col)].width = 18
    
    # Sheet 5: DTC Configuration (19服务配置)
    ws_dtc = wb.create_sheet("DTCConfig")
    
    dtc_headers = ["DTC(Hex)", "Name", "Severity", "FaultType", 
                   "AgingCounter", "Snapshot", "ExtendedData", "Description"]
    ws_dtc.append(dtc_headers)
    
    for col in range(1, len(dtc_headers) + 1):
        cell = ws_dtc.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    dtc_data = [
        ["0x0100", "MAF_Circuit", 2, "Circuit", 40, "Yes", "Yes", "空气质量流量计电路故障"],
        ["0x0101", "MAF_Performance", 2, "Performance", 40, "Yes", "Yes", "空气质量流量计性能故障"],
        ["0x0300", "Misfire_Detect", 3, "Plausibility", 40, "Yes", "Yes", "随机/多缸失火"],
        ["0x1601", "ECM_CommLost", 2, "Communication", 40, "Yes", "No", "与ECM通信丢失"],
        ["0xU1601", "ABS_CommLost", 2, "Communication", 40, "Yes", "No", "与ABS通信丢失"],
    ]
    for row in dtc_data:
        ws_dtc.append(row)
    
    for col in range(1, len(dtc_headers) + 1):
        ws_dtc.column_dimensions[get_column_letter(col)].width = 18
    
    # Sheet 6: Communication Configuration
    ws_comm = wb.create_sheet("Communication")
    
    comm_headers = ["Parameter", "Value", "Unit", "Description"]
    ws_comm.append(comm_headers)
    
    for col in range(1, len(comm_headers) + 1):
        cell = ws_comm.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    comm_data = [
        ["RequestID_Functional", "0x7DF", "Hex", "功能请求ID"],
        ["RequestID_Physical", "0x735", "Hex", "物理请求ID"],
        ["ResponseID", "0x73D", "Hex", "响应ID"],
        ["CAN_Baudrate", "500", "Kbps", "CAN波特率"],
        ["TxPduId", "20", "Dec", "发送PDU ID"],
        ["RxPduId_Functional", "10", "Dec", "功能接收PDU ID"],
        ["RxPduId_Physical", "11", "Dec", "物理接收PDU ID"],
    ]
    for row in comm_data:
        ws_comm.append(row)
    
    for col in range(1, len(comm_headers) + 1):
        ws_comm.column_dimensions[get_column_letter(col)].width = 25
    
    # Save workbook
    wb.save(output_path)
    print(f"✅ Created example config: {output_path}")
    return True


# =============================================================================
# Code Generator
# =============================================================================

class UDSConfigParser:
    """Parse UDS configuration from Excel"""
    
    def __init__(self, excel_path: str):
        self.wb = openpyxl.load_workbook(excel_path, data_only=True)
        self.config = {
            "sessions": [],
            "dids": [],
            "security": [],
            "routines": [],
            "dtcs": [],
            "communication": {},
        }
    
    def parse_all(self):
        """Parse all sheets"""
        self._parse_session_config()
        self._parse_did_config()
        self._parse_security_config()
        self._parse_routine_config()
        self._parse_dtc_config()
        self._parse_communication()
        return self.config
    
    def _parse_session_config(self):
        """Parse SessionConfig sheet"""
        ws = self.wb["SessionConfig"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            session = {
                "type": int(row[0]) if isinstance(row[0], int) else int(str(row[0]), 0),
                "name": str(row[1]) if row[1] else "",
                "p2_max": int(row[2]) if row[2] else 50,
                "p2_star_max": int(row[3]) if row[3] else 5000,
                "s3_client": int(row[4]) if row[4] else 2000,
                "s3_server": int(row[5]) if row[5] else 5000,
                "services": str(row[6]).split(",") if row[6] else [],
                "description": str(row[7]) if row[7] else "",
            }
            self.config["sessions"].append(session)
    
    def _parse_did_config(self):
        """Parse DIDConfig sheet"""
        ws = self.wb["DIDConfig"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            did_val = row[0]
            if isinstance(did_val, str):
                did_val = int(did_val, 0)
            else:
                did_val = int(did_val)
            
            did = {
                "id": did_val,
                "name": str(row[1]).replace(" ", "_") if row[1] else f"DID_{did_val:04X}",
                "length": int(row[2]) if row[2] else 1,
                "readonly": str(row[3]).lower() in ["yes", "true", "1"] if row[3] else True,
                "security_level": int(row[4]) if row[4] else 0x01,
                "session_mask": int(str(row[5]), 0) if row[5] else 0x07,
                "data_type": str(row[6]) if row[6] else "HEX",
                "description": str(row[7]) if row[7] else "",
            }
            self.config["dids"].append(did)
    
    def _parse_security_config(self):
        """Parse SecurityConfig sheet"""
        ws = self.wb["SecurityConfig"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            sec = {
                "level": int(row[0]),
                "name": str(row[1]) if row[1] else f"Level{row[0]}",
                "seed_length": int(row[2]) if row[2] else 4,
                "key_length": int(row[3]) if row[3] else 4,
                "max_attempts": int(row[4]) if row[4] else 3,
                "delay_time": int(row[5]) if row[5] else 10000,
                "algorithm": str(row[6]) if row[6] else "XOR",
                "description": str(row[7]) if row[7] else "",
            }
            self.config["security"].append(sec)
    
    def _parse_routine_config(self):
        """Parse RoutineConfig sheet"""
        ws = self.wb["RoutineConfig"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            rid_val = row[0]
            if isinstance(rid_val, str):
                rid_val = int(rid_val, 0)
            else:
                rid_val = int(rid_val)
            
            routine = {
                "id": rid_val,
                "name": str(row[1]).replace(" ", "_") if row[1] else f"RID_{rid_val:04X}",
                "start": str(row[2]).lower() in ["yes", "true", "1"] if row[2] else True,
                "stop": str(row[3]).lower() in ["yes", "true", "1"] if row[3] else False,
                "result": str(row[4]).lower() in ["yes", "true", "1"] if row[4] else True,
                "security_level": int(row[5]) if row[5] else 0x01,
                "session_mask": int(str(row[6]), 0) if row[6] else 0x07,
                "description": str(row[7]) if row[7] else "",
            }
            self.config["routines"].append(routine)
    
    def _parse_dtc_config(self):
        """Parse DTCConfig sheet"""
        ws = self.wb["DTCConfig"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            dtc_val = row[0]
            if isinstance(dtc_val, str):
                dtc_val = int(dtc_val.replace("U", ""), 0)
            else:
                dtc_val = int(dtc_val)
            
            dtc = {
                "code": dtc_val,
                "name": str(row[1]).replace(" ", "_") if row[1] else f"DTC_{dtc_val:06X}",
                "severity": int(row[2]) if row[2] else 2,
                "fault_type": str(row[3]) if row[3] else "Circuit",
                "aging_counter": int(row[4]) if row[4] else 40,
                "snapshot": str(row[5]).lower() in ["yes", "true", "1"] if row[5] else True,
                "extended": str(row[6]).lower() in ["yes", "true", "1"] if row[6] else False,
                "description": str(row[7]) if row[7] else "",
            }
            self.config["dtcs"].append(dtc)
    
    def _parse_communication(self):
        """Parse Communication sheet"""
        ws = self.wb["Communication"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            key = str(row[0])
            val = row[1]
            if isinstance(val, str) and val.startswith("0x"):
                val = int(val, 0)
            elif isinstance(val, str) and val.isdigit():
                val = int(val)
            self.config["communication"][key] = val


class UDSCodeGenerator:
    """Generate C code from UDS configuration"""
    
    def __init__(self, config: dict):
        self.config = config
        self.header_lines = []
        self.source_lines = []
    
    def generate(self) -> tuple:
        """Generate header and source files"""
        self._generate_header()
        self._generate_source()
        return "\n".join(self.header_lines), "\n".join(self.source_lines)
    
    def _generate_header(self):
        """Generate header file content"""
        comm = self.config.get("communication", {})
        
        header = f"""/**
 * @file    UdsConfig_Generated.h
 * @brief   UDS Diagnostic Configuration (Auto-generated)
 * @author  UDS Generator
 * @date    {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
 * 
 * DO NOT EDIT THIS FILE MANUALLY!
 * This file is auto-generated from UDS_Config.xlsx
 */

#ifndef UDSCONFIG_GENERATED_H
#define UDSCONFIG_GENERATED_H

#ifdef __cplusplus
extern "C" {{
#endif

/*==================================================================================================
 *                              Includes
 *=================================================================================================*/
#include "Dcm.h"
#include "Std_Types.h"

/*==================================================================================================
 *                              Communication Configuration
 *=================================================================================================*/
#define UDS_CAN_REQUEST_FUNC_ID     {comm.get('RequestID_Functional', '0x7DF'):#05x}U
#define UDS_CAN_REQUEST_PHY_ID      {comm.get('RequestID_Physical', '0x735'):#05x}U
#define UDS_CAN_RESPONSE_ID         {comm.get('ResponseID', '0x73D'):#05x}U
#define UDS_CAN_BAUDRATE            {comm.get('CAN_Baudrate', 500)}000U

#define UDS_TX_PDU_ID               {comm.get('TxPduId', 20)}U
#define UDS_RX_PDU_ID_FUNC          {comm.get('RxPduId_Functional', 10)}U
#define UDS_RX_PDU_ID_PHY           {comm.get('RxPduId_Physical', 11)}U

/*==================================================================================================
 *                              Session Configuration
 *=================================================================================================*/
"""
        self.header_lines.append(header)
        
        # Session enums
        self.header_lines.append("typedef enum {")
        for session in self.config.get("sessions", []):
            self.header_lines.append(f"    UDS_SESSION_{session['name'].upper()} = 0x{session['type']:02X},")
        self.header_lines.append("    UDS_SESSION_MAX")
        self.header_lines.append("} Uds_SessionType;\n")
        
        # Session timing
        self.header_lines.append("/* Session Timing Parameters (ms) */")
        for session in self.config.get("sessions", []):
            name = session['name'].upper()
            self.header_lines.append(f"#define UDS_P2_MAX_{name:<20} {session['p2_max']}U")
            self.header_lines.append(f"#define UDS_P2STAR_MAX_{name:<17} {session['p2_star_max']}U")
            self.header_lines.append(f"#define UDS_S3_CLIENT_{name:<18} {session['s3_client']}U")
            self.header_lines.append(f"#define UDS_S3_SERVER_{name:<18} {session['s3_server']}U")
        self.header_lines.append("")
        
        # DID definitions
        self.header_lines.append("/*==================================================================================================")
        self.header_lines.append(" *                              DID Configuration")
        self.header_lines.append(" *=================================================================================================*/")
        for did in self.config.get("dids", []):
            self.header_lines.append(f"#define UDS_DID_{did['name'].upper():<20} 0x{did['id']:04X}U  /* {did['description']} */")
        self.header_lines.append("")
        
        # Security levels
        self.header_lines.append("/*==================================================================================================")
        self.header_lines.append(" *                              Security Access Configuration")
        self.header_lines.append(" *=================================================================================================*/")
        for sec in self.config.get("security", []):
            self.header_lines.append(f"#define UDS_SEC_{sec['name'].upper():<20} 0x{sec['level']:02X}U")
        self.header_lines.append("")
        
        # Routine IDs
        self.header_lines.append("/*==================================================================================================")
        self.header_lines.append(" *                              Routine Control Configuration")
        self.header_lines.append(" *=================================================================================================*/")
        for routine in self.config.get("routines", []):
            self.header_lines.append(f"#define UDS_RID_{routine['name'].upper():<20} 0x{routine['id']:04X}U  /* {routine['description']} */")
        self.header_lines.append("")
        
        # DTC definitions
        self.header_lines.append("/*==================================================================================================")
        self.header_lines.append(" *                              DTC Configuration")
        self.header_lines.append(" *=================================================================================================*/")
        for dtc in self.config.get("dtcs", []):
            self.header_lines.append(f"#define UDS_DTC_{dtc['name'].upper():<20} 0x{dtc['code']:06X}U  /* {dtc['description']} */")
        self.header_lines.append("")
        
        # Function prototypes
        self.header_lines.append("/*==================================================================================================")
        self.header_lines.append(" *                              Function Prototypes")
        self.header_lines.append(" *=================================================================================================*/")
        self.header_lines.append("void UdsConfig_Init(void);")
        self.header_lines.append("void UdsConfig_MainFunction(void);")
        self.header_lines.append("Std_ReturnType UdsConfig_ReadDid(uint16 did, uint8* data, uint16* len);")
        self.header_lines.append("Std_ReturnType UdsConfig_WriteDid(uint16 did, const uint8* data, uint16 len);")
        self.header_lines.append("")
        
        # End of header
        self.header_lines.append("#ifdef __cplusplus")
        self.header_lines.append("}")
        self.header_lines.append("#endif")
        self.header_lines.append("")
        self.header_lines.append("#endif /* UDSCONFIG_GENERATED_H */")
    
    def _generate_source(self):
        """Generate source file content"""
        source = f"""/**
 * @file    UdsConfig_Generated.c
 * @brief   UDS Diagnostic Configuration Implementation (Auto-generated)
 * @author  UDS Generator
 * @date    {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
 * 
 * DO NOT EDIT THIS FILE MANUALLY!
 * This file is auto-generated from UDS_Config.xlsx
 */

#include "UdsConfig_Generated.h"
#include <string.h>

/*==================================================================================================
 *                              Internal Data Storage
 *=================================================================================================*/
"""
        self.source_lines.append(source)
        
        # DID data storage
        self.source_lines.append("/* DID Data Storage */")
        for did in self.config.get("dids", []):
            size = did['length']
            name = did['name']
            if did['data_type'] == "ASCII":
                # Create byte array from string
                default_str = name[:size].ljust(size)
                hex_bytes = ", ".join([f"0x{ord(c):02X}" for c in default_str])
                self.source_lines.append(f"static uint8 Uds_DidData_{name}[{size}] = {{{hex_bytes}}};")
            else:
                self.source_lines.append(f"static uint8 Uds_DidData_{name}[{size}] = {{{', '.join(['0x00'] * size)}}};")
        self.source_lines.append("")
        
        # Session table
        self.source_lines.append("/* Session Configuration Table */")
        self.source_lines.append("typedef struct {")
        self.source_lines.append("    uint8  sessionType;")
        self.source_lines.append("    uint16 p2Max;")
        self.source_lines.append("    uint16 p2StarMax;")
        self.source_lines.append("    uint16 s3Client;")
        self.source_lines.append("    uint16 s3Server;")
        self.source_lines.append("    uint32 supportedServices;")
        self.source_lines.append("} Uds_SessionConfigType;")
        self.source_lines.append("")
        self.source_lines.append("static const Uds_SessionConfigType Uds_SessionTable[] = {")
        for session in self.config.get("sessions", []):
            # Convert service list to bitmask
            service_bits = []
            for s in session['services']:
                s = s.strip()
                if s.startswith('0x'):
                    sid = int(s, 16)
                else:
                    sid = int(s, 16) if all(c in '0123456789ABCDEFabcdef' for c in s) else int(s)
                service_bits.append(f"(1UL<<0x{sid:02X})")
            services = " | ".join(service_bits) if service_bits else "0"
            self.source_lines.append(f"    {{ /* {session['name']} */")
            self.source_lines.append(f"        .sessionType = 0x{session['type']:02X},")
            self.source_lines.append(f"        .p2Max = {session['p2_max']},")
            self.source_lines.append(f"        .p2StarMax = {session['p2_star_max']},")
            self.source_lines.append(f"        .s3Client = {session['s3_client']},")
            self.source_lines.append(f"        .s3Server = {session['s3_server']},")
            self.source_lines.append(f"        .supportedServices = {services}")
            self.source_lines.append(f"    }},")
        self.source_lines.append("};")
        self.source_lines.append("")
        
        # DID table - match Dcm_DidInfoType structure
        self.source_lines.append("/* DID Configuration Table - Compatible with Dcm_DidInfoType */")
        self.source_lines.append("typedef struct {")
        self.source_lines.append("    uint16  Did;")
        self.source_lines.append("    uint8*  Data;")
        self.source_lines.append("    uint16  DataLength;")
        self.source_lines.append("    boolean IsReadOnly;")
        self.source_lines.append("} Uds_DidInfoType;")
        self.source_lines.append("")
        self.source_lines.append("/* Additional DID properties */")
        self.source_lines.append("typedef struct {")
        self.source_lines.append("    uint16  Did;")
        self.source_lines.append("    uint8   securityLevel;")
        self.source_lines.append("    uint8   sessionMask;")
        self.source_lines.append("} Uds_DidPropType;")
        self.source_lines.append("")
        self.source_lines.append("static const Uds_DidInfoType Uds_DidTable[] = {")
        for did in self.config.get("dids", []):
            self.source_lines.append(f"    {{ /* 0x{did['id']:04X} - {did['description']} */")
            self.source_lines.append(f"        .Did = 0x{did['id']:04X},")
            self.source_lines.append(f"        .Data = Uds_DidData_{did['name']},")
            self.source_lines.append(f"        .DataLength = {did['length']},")
            self.source_lines.append(f"        .IsReadOnly = {str(did['readonly']).upper()}")
            self.source_lines.append(f"    }},")
        self.source_lines.append("};")
        self.source_lines.append("")
        self.source_lines.append("static const Uds_DidPropType Uds_DidProps[] = {")
        for did in self.config.get("dids", []):
            self.source_lines.append(f"    {{ /* 0x{did['id']:04X} */")
            self.source_lines.append(f"        .Did = 0x{did['id']:04X},")
            self.source_lines.append(f"        .securityLevel = 0x{did['security_level']:02X},")
            self.source_lines.append(f"        .sessionMask = 0x{did['session_mask']:02X}")
            self.source_lines.append(f"    }},")
        self.source_lines.append("};")
        self.source_lines.append("")
        
        # DTC table - match Dcm_DtcInfoType structure
        self.source_lines.append("/* DTC Configuration Table - Compatible with Dcm_DtcInfoType */")
        self.source_lines.append("static Dcm_DtcInfoType Uds_DtcTable[] = {")
        for dtc in self.config.get("dtcs", []):
            self.source_lines.append(f"    {{ /* 0x{dtc['code']:06X} - {dtc['description']} */")
            self.source_lines.append(f"        .Dtc = 0x{dtc['code']:06X}U,")
            self.source_lines.append(f"        .Status = 0x00,")
            self.source_lines.append(f"        .Snapshot = {{0}},")
            self.source_lines.append(f"        .ExtData = {{0}}")
            self.source_lines.append(f"    }},")
        self.source_lines.append("};")
        self.source_lines.append("")
        
        # Functions
        self.source_lines.append("/*==================================================================================================")
        self.source_lines.append(" *                              Public Functions")
        self.source_lines.append(" *=================================================================================================*/")
        self.source_lines.append("")
        
        # Read DID function
        self.source_lines.append("Std_ReturnType UdsConfig_ReadDid(uint16 did, uint8* data, uint16* len)")
        self.source_lines.append("{")
        self.source_lines.append("    uint8 i;")
        self.source_lines.append("    for (i = 0; i < sizeof(Uds_DidTable)/sizeof(Uds_DidTable[0]); i++) {")
        self.source_lines.append("        if (Uds_DidTable[i].Did == did) {")
        self.source_lines.append("            if (*len >= Uds_DidTable[i].DataLength) {")
        self.source_lines.append("                memcpy(data, Uds_DidTable[i].Data, Uds_DidTable[i].DataLength);")
        self.source_lines.append("                *len = Uds_DidTable[i].DataLength;")
        self.source_lines.append("                return E_OK;")
        self.source_lines.append("            }")
        self.source_lines.append("            return E_NOT_OK;")
        self.source_lines.append("        }")
        self.source_lines.append("    }")
        self.source_lines.append("    return E_NOT_OK;")
        self.source_lines.append("}")
        self.source_lines.append("")
        
        # Write DID function
        self.source_lines.append("Std_ReturnType UdsConfig_WriteDid(uint16 did, const uint8* data, uint16 len)")
        self.source_lines.append("{")
        self.source_lines.append("    uint8 i;")
        self.source_lines.append("    for (i = 0; i < sizeof(Uds_DidTable)/sizeof(Uds_DidTable[0]); i++) {")
        self.source_lines.append("        if (Uds_DidTable[i].Did == did) {")
        self.source_lines.append("            if (Uds_DidTable[i].IsReadOnly) {")
        self.source_lines.append("                return E_NOT_OK; /* Read-only DID */")
        self.source_lines.append("            }")
        self.source_lines.append("            if (len == Uds_DidTable[i].DataLength) {")
        self.source_lines.append("                memcpy(Uds_DidTable[i].Data, data, len);")
        self.source_lines.append("                return E_OK;")
        self.source_lines.append("            }")
        self.source_lines.append("            return E_NOT_OK;")
        self.source_lines.append("        }")
        self.source_lines.append("    }")
        self.source_lines.append("    return E_NOT_OK;")
        self.source_lines.append("}")
        self.source_lines.append("")
        
        # Init function
        self.source_lines.append("void UdsConfig_Init(void)")
        self.source_lines.append("{")
        self.source_lines.append("    /* Initialize DID default values */")
        for did in self.config.get("dids", []):
            if did['data_type'] == "ASCII":
                name = did['name']
                default = name[:did['length']].ljust(did['length'])
                hex_vals = ", ".join([f"0x{ord(c):02X}" for c in default])
                self.source_lines.append(f"    /* {did['name']}: {default} */")
        self.source_lines.append("    /* Configuration loaded successfully */")
        self.source_lines.append("}")


def generate_makefile_rules(output_dir: str):
    """Generate Makefile addition instructions"""
    makefile_addition = """
# UDS Configuration Files (Auto-generated)
UDS_CFG_SRC = $(AUTOSAR_DIR)/UdsConfig/UdsConfig_Generated.c
UDS_CFG_INC = $(AUTOSAR_DIR)/UdsConfig

# Add to sources
C_SOURCES += $(UDS_CFG_SRC)

# Add to include paths
C_INCLUDES += -I$(UDS_CFG_INC)
"""
    return makefile_addition


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        print("Usage:")
        print(f"    python3 {sys.argv[0]} <config.xlsx> [output_dir]")
        print(f"    python3 {sys.argv[0]} --example")
        sys.exit(1)
    
    if sys.argv[1] == "--example":
        # Create example configuration
        output_path = "UDS_Config_Example.xlsx"
        create_example_config(output_path)
        print(f"\\nExample configuration created: {output_path}")
        print("You can edit this file and then run:")
        print(f"    python3 {sys.argv[0]} {output_path}")
        sys.exit(0)
    
    excel_path = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else "generated"
    
    if not os.path.exists(excel_path):
        print(f"Error: File not found: {excel_path}")
        print(f"Run 'python3 {sys.argv[0]} --example' to create an example config")
        sys.exit(1)
    
    # Parse configuration
    print(f"Parsing configuration: {excel_path}")
    parser = UDSConfigParser(excel_path)
    config = parser.parse_all()
    
    print(f"  - Sessions: {len(config['sessions'])}")
    print(f"  - DIDs: {len(config['dids'])}")
    print(f"  - Security levels: {len(config['security'])}")
    print(f"  - Routines: {len(config['routines'])}")
    print(f"  - DTCs: {len(config['dtcs'])}")
    
    # Generate code
    print(f"\\nGenerating code to: {output_dir}/")
    os.makedirs(output_dir, exist_ok=True)
    
    generator = UDSCodeGenerator(config)
    header_content, source_content = generator.generate()
    
    # Write files
    header_path = os.path.join(output_dir, "UdsConfig_Generated.h")
    source_path = os.path.join(output_dir, "UdsConfig_Generated.c")
    
    with open(header_path, "w") as f:
        f.write(header_content)
    print(f"  ✓ {header_path}")
    
    with open(source_path, "w") as f:
        f.write(source_content)
    print(f"  ✓ {source_path}")
    
    # Generate integration guide
    guide_path = os.path.join(output_dir, "INTEGRATION.md")
    guide_content = f"""# UDS Configuration Integration Guide

## Generated Files

- `UdsConfig_Generated.h` - Header file with all UDS configuration macros
- `UdsConfig_Generated.c` - Implementation with data tables and functions

## Quick Integration

### 1. Copy Files

```bash
cp {output_dir}/* $YOUR_PROJECT/AUTOSAR/UdsConfig/
```

### 2. Update Makefile

Add the generated source file to your build:

```makefile
# In your Makefile
C_SOURCES += AUTOSAR/UdsConfig/UdsConfig_Generated.c
C_INCLUDES += -IAUTOSAR/UdsConfig
```

### 3. Update DCM Configuration

Modify your `AUTOSAR_Cfg.h` to use generated configurations:

```c
#include "UdsConfig_Generated.h"

/* Replace static DCM config with generated ones */
#define DCM_NUM_DIDS    (sizeof(Uds_DidTable)/sizeof(Uds_DidTable[0]))
#define DCM_NUM_DTCS    (sizeof(Uds_DtcTable)/sizeof(Uds_DtcTable[0]))

/* Use generated read/write functions */
#define DCM_READ_DID_FUNC   UdsConfig_ReadDid
#define DCM_WRITE_DID_FUNC  UdsConfig_WriteDid
```

### 4. Initialize

Call initialization in your main:

```c
#include "UdsConfig_Generated.h"

void main(void) {{
    /* ... other init ... */
    UdsConfig_Init();
    Dcm_Init(&Dcm_Config);
    /* ... */
}}
```

## Configuration Summary

| Parameter | Value |
|-----------|-------|
| Functional Request ID | {config['communication'].get('RequestID_Functional', '0x7DF'):#x} |
| Physical Request ID | {config['communication'].get('RequestID_Physical', '0x735'):#x} |
| Response ID | {config['communication'].get('ResponseID', '0x73D'):#x} |
| Sessions | {len(config['sessions'])} |
| DIDs | {len(config['dids'])} |
| DTCs | {len(config['dtcs'])} |

## Regenerating Code

After modifying `{excel_path}`:

```bash
python3 tools/uds_generator/uds_generator.py {excel_path} {output_dir}
```
"""
    with open(guide_path, "w") as f:
        f.write(guide_content)
    print(f"  ✓ {guide_path}")
    
    print("\\n✅ Code generation complete!")
    print(f"\\nNext steps:")
    print(f"  1. Review generated files in: {output_dir}/")
    print(f"  2. Follow INTEGRATION.md for setup instructions")
    print(f"  3. Rebuild your project")


if __name__ == "__main__":
    main()
